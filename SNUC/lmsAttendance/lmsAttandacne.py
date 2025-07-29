from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
import time
import argparse
import getpass
# Set download directory path (absolute path)
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import re
import sys
import subprocess
import importlib
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
import glob
##########################################################################
# Parse command-line arguments
parser = argparse.ArgumentParser(description='Automate LMS attendance export.')

parser.add_argument('-u', '--username', required=True, help='LMS Username')
#parser.add_argument('-p', '--password', required=True, help='LMS Password')
parser.add_argument('-c', '--course', required=True, help='Course name link text')
parser.add_argument("-l", "--length", type=int, required=True, help="Number of Lectures")

args = parser.parse_args()

USERNAME = args.username
#PASSWORD = args.password
PASSWORD = getpass.getpass("🔒 Enter LMS Password: ")
COURSE_NAME = args.course
l = args.length
required_packages = {
    "selenium": "selenium",
    "webdriver_manager": "webdriver-manager",
    "openpyxl": "openpyxl",
    "docx": "python-docx",        # 'docx' is installed via 'python-docx'
    "reportlab": "reportlab"
}
#######################################################################################
def rename_single_xlsx(new_filename):
    if not new_filename.endswith('.xlsx'):
        new_filename += '.xlsx'

    current_dir = os.getcwd()  # Changed from __file__ to current working directory
    print("Looking in:", current_dir)

    xlsx_files = [f for f in os.listdir(current_dir) if f.endswith('.xlsx')]
    print("Found files:", xlsx_files)

    if len(xlsx_files) == 1:
        old_filename = xlsx_files[0]
        os.rename(old_filename, new_filename)
        print(f"Renamed '{old_filename}' to '{new_filename}'")
    elif len(xlsx_files) == 0:
        print("❌ No .xlsx file found in the current directory.")
    else:
        print("⚠️ Multiple .xlsx files found. Rename skipped.")
##########################################################################

def remove_first_3_rows_from_xlsx():
    # Get current working directory
    current_dir = os.getcwd()
    
    # Find the first .xlsx file
    xlsx_files = [f for f in os.listdir(current_dir) if f.endswith('.xlsx')]
    
    if not xlsx_files:
        print("❌ No .xlsx file found in the current directory.")
        return
    elif len(xlsx_files) > 1:
        print("⚠️ Multiple .xlsx files found. Edit script to handle specific one.")
        return

    file_path = os.path.join(current_dir, xlsx_files[0])
    
    # Load workbook and active sheet
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Delete the first 3 rows
    ws.delete_rows(1, 4)  # Delete from row 1, total 3 rows
    
    # Save it (overwrite)
    wb.save(file_path)
    print(f"✅ First 4 rows removed from: {xlsx_files[0]}")
##########################################################################
def export_columns_to_pdf(n, m, file, output_pdf="output.pdf"):
    if len(m) != n:
        print(f"❌ Error: n={n} but m contains {len(m)} elements: {m}")
        return

    if not os.path.exists(file):
        print("❌ File not found:", file)
        return

    # Load workbook
    wb = load_workbook(file)
    ws = wb.active

    # Extract selected columns
    data = []
    for row in ws.iter_rows(values_only=True):
        row_data = []
        for col_num in m:
            try:
                value = row[col_num - 1]
            except IndexError:
                value = ''
            row_data.append(str(value) if value is not None else '')
        data.append(row_data)

    # Create PDF
    c = canvas.Canvas(output_pdf, pagesize=A4)
    width, height = A4

    x = 0 * mm
    y = height - 20 * mm
    row_height = 7 * mm

    col_widths = [46.66 * mm] + [5.83 * mm] * (n - 1)

    row_count = 0  # Counter to track rows per page

    for row_data in data:
        current_x = x
        for i in range(n):
            c.rect(current_x, y - row_height, col_widths[i], row_height)
            c.drawString(current_x + 1 * mm, y - row_height + 2 * mm, row_data[i])
            current_x += col_widths[i]

        y -= row_height
        row_count += 1

        # Start a new page after every 25 rows
        if row_count % 25 == 0:
            c.showPage()
            y = height - 20 * mm

    c.save()
    print(f"✅ PDF saved as: {output_pdf}")

   
###########################################################################

def merge_columns(filename, output_filename=None, col1=1, col2=2, target_col=3):
    """
    Merges values from two columns in an Excel file and writes the merged result into a new column.

    Parameters:
        filename (str): Path to the input .xlsx file.
        output_filename (str): Name of the output file (if None, overwrites the input file).
        col1 (int): First column index to merge (1-based).
        col2 (int): Second column index to merge (1-based).
        target_col (int): Target column index to write merged results (1-based).
    """
    wb = load_workbook(filename)
    ws = wb.active

    for row in ws.iter_rows(min_row=1, max_col=max(col1, col2)):
        val1 = row[col1 - 1].value
        val2 = row[col2 - 1].value
        merged = f"{val1} {val2}" if val1 and val2 else val1 or val2
        ws.cell(row=row[0].row, column=target_col).value = merged

    output_path = output_filename if output_filename else filename
    wb.save(output_path)
    print(f"Merged columns saved to: {output_path}")
##########################################################################
def remove_pattern_from_excel(filename, pattern):
    """
    Removes a specific regex pattern from all cells in an Excel file.

    Parameters:
        filename (str): Path to the .xlsx file to modify.
        pattern (str): Regex pattern to remove from cell contents.
    """
    wb = load_workbook(filename)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cleaned = re.sub(pattern, '', cell.value)
                cell.value = cleaned

    wb.save(filename)
    print(f"Pattern '{pattern}' removed and saved in: {filename}")
##########################################################################
for module_name, pip_name in required_packages.items():
    try:
        importlib.import_module(module_name)
    except ImportError:
        print(f"📦 Installing missing package: {pip_name}")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name])
# Convert relative path to absolute path
download_dir = os.path.abspath("./")  # Ensures absolute path
# Clean up existing .pdf and .xlsx files
for file in glob.glob("log_book.pdf") + glob.glob("merged_output.xlsx")  + glob.glob("final_report.xlsx"):
    try:
        os.remove(file)
        print(f"🗑️ Removed: {file}")
    except Exception as e:
        print(f"❌ Failed to remove {file}: {e}")
        
#####################################################################        
chrome_options = Options()
chrome_options.add_experimental_option("prefs", {
    "download.default_directory": download_dir,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True
})

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

driver.get('https://lms.snuchennai.edu.in/login/index.php')

# Launch Chrome with the configured options

#driver.maximize_window()
# Perform actions...
username_input = driver.find_element(By.ID, "username")
username_input.send_keys(USERNAME)
time.sleep(5)
password_field = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.ID, "password"))
)
password_field.click()
password_field.clear()
password_field.send_keys(PASSWORD)
#time.sleep(5)
try:
    button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "loginbtn"))
    )
    button.click()
except Exception as e:
    print(f"Error: {e}")
    
my_courses = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.LINK_TEXT, "My courses"))
)
my_courses.click()


my_courses = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.LINK_TEXT, COURSE_NAME))
)
my_courses.click()

my_courses = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.LINK_TEXT, "Attendance"))
)
my_courses.click()

my_courses = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.LINK_TEXT, "Export"))
)
my_courses.click()


try:
    button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "id_submitbutton"))
    )
    button.click()
except Exception as e:
    print(f"Error: {e}")
    
time.sleep(5)    
rename_single_xlsx("final_report.xlsx")
remove_first_3_rows_from_xlsx()
merge_columns("final_report.xlsx", output_filename="merged_output.xlsx")
remove_pattern_from_excel("merged_output.xlsx", r"\(2/2\)")
time.sleep(2)
remove_pattern_from_excel("merged_output.xlsx", r"\?")
remove_pattern_from_excel("merged_output.xlsx", "All students")
#export_columns_to_word(12, [4, 3, 6,7,8,9,10,11,12,13,14,15], "merged_output.xlsx")
#export_columns_to_word(11, [3, 6,7,8,9,10,11,12,13,14,15], "merged_output.xlsx")

a = list(range(6, 6 + l))  # This gives [6, 7, 8, ..., 6+l-1]
m = [3] + a

export_columns_to_pdf(
    n=len(m),               # Make sure `n` matches the length of `m`
    m=m,
    file="merged_output.xlsx",
    output_pdf="log_book.pdf"
)
time.sleep(5)
driver.quit()
